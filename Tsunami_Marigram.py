#!/usr/bin/env python3
"""
WaveSource — Google Drive Marigram OCR -> Excel (Human-in-the-loop)
==================================================================

What this script does
1) Pull marigram images from Google Drive folders (recursively)
2) OCR each image (Tesseract + a few OpenCV preprocessing variants)
3) Parse key fields (COUNTRY / STATE / LOCATION / RECORDED_DATE / SCALE)
4) Validate (but DO NOT guess) against NOAA descriptor allow-lists:
   - COUNTRY: https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/countries
   - STATE:   https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/states
   - LOCATION:https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/locations (paginated)
5) REGION_CODE (NCEI region code) is strict:
   - Only accept if explicit 2-digit code is found in OCR text AND exists in NOAA regions list:
     https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/regions
6) LOCATION_SHORT (IOC station code) is strict:
   - Scraped from IOC station list page and resolved by exact (COUNTRY, LOCATION) match:
     https://www.ioc-sealevelmonitoring.org/list.php
   - If not found: leave blank (human fills)
7) Optional: geocode LAT/LON from (LOCATION, STATE, COUNTRY) using Nominatim
8) Human-in-the-loop review:
   - If a field is missing or fails allow-list matching, we get a quick CLI prompt to accept/edit
9) Append structured rows to Excel (.xlsx), and keep a progress log so it can be resumed

Install (pip)
  pip install opencv-python pillow pytesseract openpyxl numpy requests beautifulsoup4 \
              google-api-python-client google-auth-httplib2 google-auth-oauthlib geopy

System deps
  - Tesseract binary installed and on PATH

Google Drive Auth
  - Create OAuth Client credentials (Desktop) in Google Cloud Console
  - Save as ./credentials.json
  - First run creates ./token.json after browser auth
"""

from __future__ import annotations

import argparse
import io
import json
import random
import re
import sys
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import cv2  # type: ignore
import numpy as np  # type: ignore
import pytesseract  # type: ignore
import requests  # type: ignore
from PIL import Image  # type: ignore
from bs4 import BeautifulSoup  # type: ignore

# Excel appending without re-reading the whole file each time
from openpyxl import Workbook, load_workbook  # type: ignore

# Google Drive API
from googleapiclient.discovery import build  # type: ignore
from googleapiclient.errors import HttpError  # type: ignore
from googleapiclient.http import MediaIoBaseDownload  # type: ignore
from google.oauth2.credentials import Credentials  # type: ignore
from google_auth_oauthlib.flow import InstalledAppFlow  # type: ignore
from google.auth.transport.requests import Request  # type: ignore

# geocoding
try:
    from geopy.geocoders import Nominatim  # type: ignore
    from geopy.extra.rate_limiter import RateLimiter  # type: ignore
except Exception:
    Nominatim = None
    RateLimiter = None


# ---------------------------
# Output columns
# ---------------------------
DEFAULT_COLUMNS = [
    "FILE_NAME", "COUNTRY", "STATE", "LOCATION", "LOCATION_SHORT", "REGION_CODE",
    "START_RECORD", "END_RECORD", "TSEVENT_ID", "TSRUNUP_ID", "RECORDED_DATE",
    "LATITUDE", "LONGITUDE", "IMAGES", "SCALE", "MICROFILM_NAME", "COMMENTS",
]

# ---------------------------
# NOAA descriptor endpoints
# ---------------------------
NOAA_COUNTRIES_URL = "https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/countries?itemsPerPage=200&page=1"
NOAA_STATES_URL    = "https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/states?itemsPerPage=200&page=1"
NOAA_REGIONS_URL   = "https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/regions?itemsPerPage=200&page=1"
NOAA_LOCATIONS_URL = "https://www.ngdc.noaa.gov/hazel/hazard-service/api/v1/descriptors/tsunamis/marigrams/locations?itemsPerPage=200&page={page}"

# IOC station list (LOCATION_SHORT)
IOC_LIST_URL = "https://www.ioc-sealevelmonitoring.org/list.php"

# ---------------------------
# Regexes
# ---------------------------
DATE_PATTERNS = [
    re.compile(r"(?<!\d)(?P<y>19\d{2}|20\d{2})-(?P<m>0[1-9]|1[0-2])-(?P<d>0[1-9]|[12]\d|3[01])(?!\d)"),
    re.compile(r"(?<!\d)(?P<m>0?[1-9]|1[0-2])[\-/](?P<d>0?[1-9]|[12]\d|3[01])[\-/](?P<y>19\d{2}|20\d{2})(?!\d)"),
    re.compile(r"(?<!\w)(?P<d>0?[1-9]|[12]\d|3[01])\s+(?P<mon>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)[a-z]*\s+(?P<y>19\d{2}|20\d{2})(?!\w)", re.I),
]
MONTH_MAP = {'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06','JUL':'07','AUG':'08','SEP':'09','SEPT':'09','OCT':'10','NOV':'11','DEC':'12'}

SCALE_PATTERNS = [
    re.compile(r"(?:SCALE\s*[:=]?\s*)?1\s*[:/]\s*(?P<den>\d{1,4})", re.I),
]

# UPPERCASE triple split on 2+ spaces
UPPER_TRIPLE_SPLIT = re.compile(
    r"^([A-Z][A-Z\- .'()&/]+?)\s{2,}([A-Z][A-Z\- .'()&/]+?)\s{2,}([A-Z0-9][A-Z0-9\- .,'()&/]+)$"
)

# Strict IOC code appearance (4-5 usually, allow 3-6 to be safe)
IOC_CODE_RE = re.compile(r"^[A-Za-z0-9]{3,6}$")

# Quick “anchors” that are common on these sheets
OCR_ANCHORS = [
    r"\bCOUNTRY\b",
    r"\bSTATE\b",
    r"\bLOCATION\b",
    r"\bSCALE\b",
    r"\bREGION\b",
]


# ---------------------------
# Data model
# ---------------------------
@dataclass
class Row:
    FILE_NAME: str
    COUNTRY: str = ""
    STATE: str = ""
    LOCATION: str = ""
    LOCATION_SHORT: str = ""
    REGION_CODE: str = ""
    START_RECORD: str = ""
    END_RECORD: str = ""
    TSEVENT_ID: str = ""
    TSRUNUP_ID: str = ""
    RECORDED_DATE: str = ""
    LATITUDE: str = ""
    LONGITUDE: str = ""
    IMAGES: str = ""
    SCALE: str = ""
    MICROFILM_NAME: str = ""
    COMMENTS: str = ""


# ---------------------------
# Small utils
# ---------------------------
def _upper(s: str) -> str:
    return (s or "").strip().upper()

def sanitize_text(text: str) -> str:
    text = text.replace("\x0c", " ")
    text = re.sub(r"[\u200b\u200c\u200d]", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text

def safe_filename(name: str) -> str:
    name = re.sub(r"[^\w.\- ]+", "_", name)
    return name.strip() or "file"

def is_image_name(name: str) -> bool:
    ext = Path(name).suffix.lower()
    return ext in {".tif", ".tiff", ".png", ".jpg", ".jpeg", ".webp"}

def ensure_excel(path: str) -> None:
    p = Path(path)
    if p.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(DEFAULT_COLUMNS)
    wb.save(path)

def append_rows_to_excel(path: str, rows: List[Row]) -> None:
    """
    Appends rows without loading the entire sheet into memory each time.
    This stays fast even when you’re writing thousands of images.
    """
    ensure_excel(path)
    wb = load_workbook(path)
    ws = wb.active

    # If someone created a blank workbook manually, fix header.
    if ws.max_row == 1 and all((ws.cell(1, i+1).value or "") == "" for i in range(len(DEFAULT_COLUMNS))):
        ws.delete_rows(1, 1)
        ws.append(DEFAULT_COLUMNS)

    for r in rows:
        d = asdict(r)
        ws.append([d.get(col, "") for col in DEFAULT_COLUMNS])

    wb.save(path)


# ---------------------------
# NOAA allow-lists + region map
# ---------------------------
def _fetch_json(url: str, timeout: int = 30, retries: int = 4, backoff: float = 1.7) -> dict:
    last_err: Optional[Exception] = None
    for i in range(retries):
        try:
            r = requests.get(url, timeout=timeout)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_err = e
            time.sleep(backoff ** i)
    raise RuntimeError(f"Failed to fetch JSON after {retries} tries: {url} :: {last_err}")

def fetch_noaa_lists() -> Tuple[Set[str], Set[str], Set[str], Dict[str, str]]:
    """
    Returns:
      countries_set, states_set, locations_set, regions_map(code->description)
    """
    countries_j = _fetch_json(NOAA_COUNTRIES_URL)
    states_j    = _fetch_json(NOAA_STATES_URL)
    regions_j   = _fetch_json(NOAA_REGIONS_URL)

    countries = {_upper(x["description"]) for x in countries_j.get("items", [])}
    states    = {_upper(x["description"]) for x in states_j.get("items", [])}
    regions   = {str(x["id"]).strip(): str(x["description"]).strip() for x in regions_j.get("items", [])}

    # Locations are paginated; don’t assume how many pages.
    page1 = _fetch_json(NOAA_LOCATIONS_URL.format(page=1))
    total_pages = int(page1.get("totalPages", 1))
    locations: Set[str] = {_upper(x["description"]) for x in page1.get("items", [])}
    for p in range(2, total_pages + 1):
        jp = _fetch_json(NOAA_LOCATIONS_URL.format(page=p))
        locations |= {_upper(x["description"]) for x in jp.get("items", [])}

    return countries, states, locations, regions

def validate_against_allow_list(value: str, allow: Set[str]) -> Tuple[str, bool]:
    """
    Returns (kept_value, needs_review).
    - If exact upper-case match exists => returns normalized UPPER value, needs_review=False
    - Else keep OCR text as-is (no guessing), needs_review=True
    """
    if not value or not value.strip():
        return "", True
    v = _upper(value)
    if v in allow:
        return v, False
    return value.strip(), True

def parse_region_code_strict(ocr_text: str, regions_map: Dict[str, str]) -> Tuple[str, bool]:
    """
    Strict: accept only explicit 2-digit code present in text AND exists in NOAA regions list.
    """
    if not ocr_text:
        return "", True

    m = re.search(r"\[(\d{2})\]", ocr_text)
    if m and m.group(1) in regions_map:
        return m.group(1), False

    m = re.search(r"\bREGION\b[^0-9]*(\d{2})\b", ocr_text, re.I)
    if m and m.group(1) in regions_map:
        return m.group(1), False

    return "", True


# ---------------------------
# IOC station code index (LOCATION_SHORT)
# ---------------------------
def fetch_ioc_station_index(timeout: int = 30, cache_path: Optional[Path] = None) -> Dict[Tuple[str, str], str]:
    """
    Scrape IOC list.php to build:
      (COUNTRY_UPPER, LOCATION_UPPER) -> IOC_CODE

    If cache_path is set, store/read IOC HTML so repeated runs are reproducible and faster.
    """
    html: str
    if cache_path and cache_path.exists():
        html = cache_path.read_text(encoding="utf-8", errors="ignore")
    else:
        r = requests.get(IOC_LIST_URL, timeout=timeout)
        r.raise_for_status()
        html = r.text
        if cache_path:
            cache_path.parent.mkdir(parents=True, exist_ok=True)
            cache_path.write_text(html, encoding="utf-8")

    soup = BeautifulSoup(html, "html.parser")

    tables = soup.find_all("table")
    target = None
    headers: List[str] = []

    for t in tables:
        ths = t.find_all("th")
        if not ths:
            continue
        header_text = " ".join(_upper(th.get_text(" ", strip=True)) for th in ths)
        if "CODE" in header_text and "COUNTRY" in header_text and "LOCATION" in header_text:
            target = t
            headers = [_upper(th.get_text(" ", strip=True)) for th in ths]
            break

    if target is None:
        raise RuntimeError("Could not find IOC station list table. IOC page markup may have changed.")

    def idx(name: str) -> int:
        for i, h in enumerate(headers):
            if h == name:
                return i
        for i, h in enumerate(headers):
            if name in h:
                return i
        raise RuntimeError(f"Could not locate IOC column: {name}")

    i_code = idx("CODE")
    i_country = idx("COUNTRY")
    i_location = idx("LOCATION")

    index: Dict[Tuple[str, str], str] = {}
    for tr in target.find_all("tr"):
        tds = tr.find_all("td")
        if not tds or len(tds) <= max(i_code, i_country, i_location):
            continue

        code = tds[i_code].get_text(" ", strip=True).strip()
        country = tds[i_country].get_text(" ", strip=True).strip()
        location = tds[i_location].get_text(" ", strip=True).strip()

        if not code:
            continue
        key = (_upper(country), _upper(location))
        if key not in index:
            index[key] = code

    return index

def resolve_location_short_strict(country: str, location: str, ioc_index: Dict[Tuple[str, str], str]) -> Tuple[str, bool]:
    if not ioc_index or not country or not location:
        return "", True
    code = ioc_index.get((_upper(country), _upper(location)), "")
    if code and IOC_CODE_RE.match(code):
        return code, False
    return "", True


# ---------------------------
# OCR pipeline
# ---------------------------
def load_image_cv(path: str) -> np.ndarray:
    img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        raise RuntimeError(f"Failed to read image: {path}")
    return img

def preprocess_variants(img: np.ndarray) -> List[np.ndarray]:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    out: List[np.ndarray] = []

    # Otsu
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    out.append(th)

    # Inverted Otsu (sometimes labels pop better)
    _, th_inv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    out.append(th_inv)

    # Adaptive threshold
    ad = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, 11)
    out.append(ad)

    # CLAHE -> Otsu
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(gray)
    _, th2 = cv2.threshold(clahe, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    out.append(th2)

    # Light blur -> Otsu (helps with speckle)
    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    _, th3 = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    out.append(th3)

    return out

def _ocr_avg_conf(pil_img: Image.Image, config: str) -> float:
    """
    Compute an average confidence score from Tesseract's image_to_data output.
    Returns 0.0 if confidence is unavailable.
    """
    confs: List[float] = []

    # Try DATAFRAME first
    try:
        df = pytesseract.image_to_data(
            pil_img,
            config=config,
            output_type=pytesseract.Output.DATAFRAME
        )
        if "conf" in df.columns:
            for c in df["conf"].tolist():
                try:
                    cf = float(c)
                    # Tesseract uses -1 for "not a word"
                    if cf >= 0:
                        confs.append(cf)
                except Exception:
                    continue
        return float(np.mean(confs)) if confs else 0.0
    except Exception:
        pass

    # Fall back to DICT
    try:
        d = pytesseract.image_to_data(
            pil_img,
            config=config,
            output_type=pytesseract.Output.DICT
        )
        for c in d.get("conf", []):
            try:
                cf = float(c)
                if cf >= 0:
                    confs.append(cf)
            except Exception:
                continue
        return float(np.mean(confs)) if confs else 0.0
    except Exception:
        return 0.0


def ocr_image(img: np.ndarray, psm: int = 6, oem: int = 3) -> Tuple[str, float]:
    """
    OCR with:
      - full text from image_to_string (better for your regex parsing)
      - avg confidence from image_to_data
    """
    config = f"--psm {psm} --oem {oem}"
    pil_img = Image.fromarray(img)

    # Full OCR text
    try:
        text_full = pytesseract.image_to_string(pil_img, config=config)
    except Exception:
        text_full = ""

    # Average confidence
    avg_conf = _ocr_avg_conf(pil_img, config=config)

    return text_full, avg_conf


def _anchor_score(text: str) -> int:
    if not text:
        return 0
    t = text.upper()
    score = 0
    for pat in OCR_ANCHORS:
        if re.search(pat, t):
            score += 1
    # Dates/scales are especially useful
    if normalize_date_to_ymd(t):
        score += 2
    if parse_scale(t):
        score += 1
    return score

def best_ocr_from_variants(img: np.ndarray, psm: int = 6, oem: int = 3) -> Tuple[str, float, np.ndarray, int]:
    """
    Pick the OCR result that looks most “structurally right”:
      1) more anchors (COUNTRY/STATE/LOCATION/etc.)
      2) then higher average confidence
      3) then longer text
    """
    best_text = ""
    best_conf = -1.0
    best_variant = img
    best_anchor = -1

    for var in preprocess_variants(img):
        text, conf = ocr_image(var, psm=psm, oem=oem)
        anc = _anchor_score(text)

        if (anc > best_anchor) or (anc == best_anchor and (conf > best_conf)) or (anc == best_anchor and conf == best_conf and len(text) > len(best_text)):
            best_text, best_conf, best_variant, best_anchor = text, conf, var, anc

    return best_text, best_conf, best_variant, best_anchor


# ---------------------------
# Parsing helpers (COUNTRY/STATE/LOCATION/DATE/SCALE)
# ---------------------------
def _looks_like_text(s: str) -> bool:
    if not s:
        return False
    # reject strings that are mostly digits/punct
    letters = sum(ch.isalpha() for ch in s)
    return letters >= 2

def parse_country_state_location(lines: List[str]) -> Tuple[str, str, str]:
    # A) Uppercase triple with 2+ spaces
    for line in lines[:20]:
        m = UPPER_TRIPLE_SPLIT.match(line.strip())
        if m:
            return m.group(1).strip(), m.group(2).strip(), m.group(3).strip()

    # B) Semicolon/comma triplets
    for line in lines[:30]:
        parts = re.split(r"\s*[;,\t]\s*", line.strip())
        if len(parts) >= 3:
            a, b, c = parts[0].strip(), parts[1].strip(), parts[2].strip()
            if _looks_like_text(a) and _looks_like_text(b) and _looks_like_text(c):
                return a, b, c

    # C) Explicit labels
    blob = "\n".join(lines[:80])
    m = re.search(r"COUNTRY[:\-\s]+([A-Z .,'()&/-]+)", blob, re.I)
    country = m.group(1).strip() if m else ""
    m = re.search(r"STATE[:\-\s]+([A-Z0-9 .,'()&/-]+)", blob, re.I)
    state = m.group(1).strip() if m else ""
    m = re.search(r"LOCATION[:\-\s]+([A-Z0-9 .,'()&/-]+)", blob, re.I)
    location = m.group(1).strip() if m else ""
    return country, state, location

def normalize_date_to_ymd(text: str) -> str:
    for pat in DATE_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        gd = {k: (v if v is None else str(v)) for k, v in m.groupdict().items()}
        if 'mon' in gd and gd['mon']:
            y = gd['y']; d = gd['d'].zfill(2)
            mon = gd['mon'].upper()[:4].replace('.', '')
            mm = MONTH_MAP.get(mon[:4], MONTH_MAP.get(mon[:3], ""))
            if y and mm and d:
                return f"{y}/{mm}/{d}"
        else:
            y = gd.get('y'); mm = gd.get('m'); d = gd.get('d')
            if y and mm and d:
                return f"{y}/{mm.zfill(2)}/{d.zfill(2)}"
    return ""

def parse_scale(text: str) -> str:
    for pat in SCALE_PATTERNS:
        m = pat.search(text)
        if m:
            return f"1:{m.group('den')}"
    return ""


# ---------------------------
# geocoding
# ---------------------------
def make_geocoder(enable: bool) -> Optional["RateLimiter"]:
    if not enable:
        return None
    if Nominatim is None or RateLimiter is None:
        print("Geocoding requested but geopy is not available. Install geopy or disable --enable-geocode.")
        return None
    geolocator = Nominatim(user_agent="wavesource_marigram_geocoder")
    return RateLimiter(geolocator.geocode, min_delay_seconds=1.0)

def geocode_latlon(country: str, state: str, location: str, geocode_fn: Optional["RateLimiter"]) -> Tuple[str, str]:
    if geocode_fn is None:
        return "", ""
    queries: List[str] = []
    if location and state and country:
        queries.append(f"{location}, {state}, {country}")
    if location and country:
        queries.append(f"{location}, {country}")
    if state and country:
        queries.append(f"{state}, {country}")
    if country:
        queries.append(country)

    for q in queries:
        try:
            loc = geocode_fn(q)
            if loc and getattr(loc, "latitude", None) is not None and getattr(loc, "longitude", None) is not None:
                return f"{float(loc.latitude):.5f}", f"{float(loc.longitude):.5f}"
        except Exception:
            continue
    return "", ""


# ---------------------------
# Google Drive helpers (with retries)
# ---------------------------
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

def drive_service() -> object:
    creds: Optional[Credentials] = None
    token_path = Path("token.json")
    cred_path = Path("credentials.json")

    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not cred_path.exists():
                raise RuntimeError("Missing credentials.json. Create OAuth client credentials and save as credentials.json.")
            flow = InstalledAppFlow.from_client_secrets_file(str(cred_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)

def _drive_call_with_retry(fn, retries: int = 5, backoff: float = 1.8):
    last_err: Optional[Exception] = None
    for i in range(retries):
        try:
            return fn()
        except HttpError as e:
            last_err = e
            # 429/5xx are common transient failures
            time.sleep(backoff ** i)
        except Exception as e:
            last_err = e
            time.sleep(backoff ** i)
    raise RuntimeError(f"Drive API call failed after {retries} tries: {last_err}")

def drive_list_children(svc: object, folder_id: str) -> List[dict]:
    """List direct children of a folder (files + subfolders)."""
    items: List[dict] = []
    page_token = None
    while True:
        def _call():
            return svc.files().list(
                q=f"'{folder_id}' in parents and trashed=false",
                fields="nextPageToken, files(id, name, mimeType)",
                pageToken=page_token,
                pageSize=1000
            ).execute()

        resp = _drive_call_with_retry(_call)
        items.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items

def drive_get_path_map(svc: object, folder_ids: List[str]) -> Dict[str, str]:
    """
    Traverse folders and build a map:
      file_id -> "Folder/Subfolder/filename.ext"
    """
    out: Dict[str, str] = {}
    stack: List[Tuple[str, str]] = [(fid, "") for fid in folder_ids]

    while stack:
        fid, prefix = stack.pop()
        children = drive_list_children(svc, fid)
        for it in children:
            name = it["name"]
            mime = it["mimeType"]
            it_id = it["id"]
            if mime == "application/vnd.google-apps.folder":
                stack.append((it_id, f"{prefix}{name}/"))
            else:
                out[it_id] = f"{prefix}{name}"
    return out

def drive_download_file(svc: object, file_id: str, dest_path: Path) -> None:
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    def _call_download():
        request = svc.files().get_media(fileId=file_id)
        with io.FileIO(str(dest_path), "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return True

    _drive_call_with_retry(_call_download)


# ---------------------------
# Human-in-the-loop review UI (CLI)
# ---------------------------
def prompt_field(label: str, current: str, needs_review: bool, allow_hint: str = "") -> str:
    flag = "!!" if needs_review else "  "
    hint = f" ({allow_hint})" if allow_hint else ""
    print(f"{flag} {label}{hint}: {current!r}")
    inp = input("    -> Enter to accept, or type correction: ").strip()
    return current if inp == "" else inp

def prompt_yes_no(msg: str, default_yes: bool = True) -> bool:
    d = "Y/n" if default_yes else "y/N"
    inp = input(f"{msg} [{d}]: ").strip().lower()
    if inp == "":
        return default_yes
    return inp in {"y", "yes"}


# ---------------------------
# Progress log (resume)
# ---------------------------
def load_processed_ids(log_path: Path) -> Set[str]:
    if not log_path.exists():
        return set()
    done: Set[str] = set()
    for line in log_path.read_text(encoding="utf-8").splitlines():
        try:
            obj = json.loads(line)
            if "file_id" in obj and obj.get("status") == "ok":
                done.add(str(obj["file_id"]))
        except Exception:
            continue
    return done

def append_log(log_path: Path, record: dict) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


# ---------------------------
# Core processing for one image
# ---------------------------
def process_one_image(
    local_path: Path,
    drive_rel_path: str,
    countries: Set[str],
    states: Set[str],
    locations: Set[str],
    regions_map: Dict[str, str],
    ioc_index: Dict[Tuple[str, str], str],
    geocode_fn: Optional["RateLimiter"],
    save_ocr_dir: Optional[Path],
    microfilm_name: str,
    microfilm_from_folder: bool,
    interactive: bool,
    psm: int,
    oem: int,
) -> Row:
    img = load_image_cv(str(local_path))
    ocr_text, conf, best_variant, anchor_score = best_ocr_from_variants(img, psm=psm, oem=oem)

    if save_ocr_dir:
        save_ocr_dir.mkdir(parents=True, exist_ok=True)
        stem = safe_filename(local_path.stem)
        (save_ocr_dir / f"{stem}.txt").write_text(ocr_text, encoding="utf-8")
        cv2.imwrite(str(save_ocr_dir / f"{stem}_best.png"), best_variant)

    text_clean = sanitize_text(ocr_text)
    lines = [ln.strip() for ln in text_clean.splitlines() if ln.strip()]

    # Parse OCR candidates
    country_ocr, state_ocr, location_ocr = parse_country_state_location(lines)
    recorded_date = normalize_date_to_ymd(text_clean)
    scale = parse_scale(text_clean)

    # Validate vs NOAA allow-lists (no guessing; keep OCR but flag)
    country, country_flag = validate_against_allow_list(country_ocr, countries)
    state, state_flag = validate_against_allow_list(state_ocr, states)
    location, location_flag = validate_against_allow_list(location_ocr, locations)

    # Region code strict
    region_code, region_flag = parse_region_code_strict(text_clean, regions_map)

    # IOC station code strict (LOCATION_SHORT)
    location_short, loc_short_flag = resolve_location_short_strict(country, location, ioc_index)

    # Lat/lon (geocode)
    lat, lon = geocode_latlon(country, state, location, geocode_fn)

    # MICROFILM_NAME from folder
    mf_name = microfilm_name
    if microfilm_from_folder:
        parts = drive_rel_path.split("/")
        if len(parts) > 1:
            mf_name = parts[0].strip() or microfilm_name

    # Comments: keep it simple and actually useful for auditing
    comments = f"avg_conf={conf:.1f}; anchors={anchor_score}; psm={psm}; oem={oem}; path={drive_rel_path}"

    row = Row(
        FILE_NAME=drive_rel_path,
        COUNTRY=country,
        STATE=state,
        LOCATION=location,
        LOCATION_SHORT=location_short,
        REGION_CODE=region_code,
        RECORDED_DATE=recorded_date,
        LATITUDE=lat,
        LONGITUDE=lon,
        IMAGES="1",
        SCALE=scale,
        MICROFILM_NAME=mf_name,
        COMMENTS=comments,
    )

    # Human-in-the-loop review if interactive
    if interactive:
        needs_any_review = any([
            country_flag, state_flag, location_flag,
            region_flag, loc_short_flag,
            (recorded_date == ""), (scale == "")
        ])

        print("\n" + "="*72)
        print(f"IMAGE: {drive_rel_path}")
        print(f"LOCAL: {local_path}")
        print("-"*72)

        if needs_any_review:
            print("Some fields need a look. Quick edit mode:\n")
        else:
            print("Looks fine. Edit anything you want:\n")

        row.COUNTRY = prompt_field("COUNTRY", row.COUNTRY, country_flag, allow_hint="NOAA allow-list")
        row.STATE = prompt_field("STATE", row.STATE, state_flag, allow_hint="NOAA allow-list")
        row.LOCATION = prompt_field("LOCATION", row.LOCATION, location_flag, allow_hint="NOAA allow-list")
        row.RECORDED_DATE = prompt_field("RECORDED_DATE (YYYY/MM/DD)", row.RECORDED_DATE, row.RECORDED_DATE == "")
        row.SCALE = prompt_field("SCALE (1:NN)", row.SCALE, row.SCALE == "")
        row.REGION_CODE = prompt_field("REGION_CODE (NCEI 2-digit)", row.REGION_CODE, region_flag, allow_hint="must exist in NOAA regions")
        row.LOCATION_SHORT = prompt_field("LOCATION_SHORT (IOC station code)", row.LOCATION_SHORT, loc_short_flag, allow_hint="from IOC list.php")
        row.LATITUDE = prompt_field("LATITUDE (decimal)", row.LATITUDE, row.LATITUDE == "", allow_hint="geocode")
        row.LONGITUDE = prompt_field("LONGITUDE (decimal)", row.LONGITUDE, row.LONGITUDE == "", allow_hint="geocode")
        row.MICROFILM_NAME = prompt_field("MICROFILM_NAME", row.MICROFILM_NAME, row.MICROFILM_NAME == "")
        row.IMAGES = prompt_field("IMAGES", row.IMAGES, row.IMAGES == "")
        row.COMMENTS = prompt_field("COMMENTS", row.COMMENTS, False)

        if not prompt_yes_no("Save this row to Excel?", default_yes=True):
            raise RuntimeError("User skipped row in review mode.")

    return row


# ---------------------------
# Main
# ---------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description="Google Drive HITL OCR marigram images -> Excel")
    ap.add_argument("--folder-ids", nargs="+", required=True, help="One or more Google Drive folder IDs")
    ap.add_argument("--out-xlsx", required=True, help="Output Excel path (.xlsx)")
    ap.add_argument("--cache-dir", default="./_drive_cache", help="Local cache directory for downloaded images")
    ap.add_argument("--save-ocr", default=None, help="Optional folder to save OCR text + best-preprocessed image")
    ap.add_argument("--resume", action="store_true", help="Resume using progress log (skip already processed files)")
    ap.add_argument("--log-path", default="./_progress/processed.jsonl", help="Progress log path (jsonl)")
    ap.add_argument("--interactive", action="store_true", help="Enable human-in-the-loop review prompts")
    ap.add_argument("--enable-geocode", action="store_true", help="Enable Nominatim geocoding for lat/lon (rate-limited)")
    ap.add_argument("--microfilm-name", default="", help="Default MICROFILM_NAME (if not using --microfilm-name-from-folder)")
    ap.add_argument("--microfilm-name-from-folder", action="store_true", help="Set MICROFILM_NAME from top-level Drive folder name")

    # Quality-of-life for testing and reproducibility
    ap.add_argument("--max-files", type=int, default=0, help="Process at most N images (0 = no limit)")
    ap.add_argument("--shuffle", action="store_true", help="Shuffle file processing order")
    ap.add_argument("--sort", action="store_true", help="Sort file processing order by path")

    # OCR knobs
    ap.add_argument("--psm", type=int, default=6, help="Tesseract page segmentation mode (default: 6)")
    ap.add_argument("--oem", type=int, default=3, help="Tesseract OCR engine mode (default: 3)")

    # IOC cache
    ap.add_argument("--ioc-cache-html", default="./_cache/ioc_list.html", help="Where to cache IOC list HTML")

    args = ap.parse_args()

    out_xlsx = str(Path(args.out_xlsx))
    cache_dir = Path(args.cache_dir)
    save_ocr_dir = Path(args.save_ocr) if args.save_ocr else None
    log_path = Path(args.log_path)
    ioc_cache_path = Path(args.ioc_cache_html) if args.ioc_cache_html else None

    # Fetch official lists
    print("Fetching NOAA allow-lists (countries/states/locations/regions)...")
    countries, states, locations, regions_map = fetch_noaa_lists()
    print(f"  countries={len(countries)}, states={len(states)}, locations={len(locations)}, regions={len(regions_map)}")

    print("Fetching IOC station list (LOCATION_SHORT codes)...")
    try:
        ioc_index = fetch_ioc_station_index(cache_path=ioc_cache_path)
        print(f"  IOC index entries={len(ioc_index)} (cache: {ioc_cache_path})")
    except Exception as e:
        print(f"  IOC fetch failed, continuing without IOC codes: {e}")
        ioc_index = {}

    geocode_fn = make_geocoder(args.enable_geocode)

    svc = drive_service()

    # Traverse drive folder(s)
    print("Listing files in Drive folders...")
    id_to_relpath = drive_get_path_map(svc, args.folder_ids)
    image_items = [(fid, rel) for fid, rel in id_to_relpath.items() if is_image_name(rel)]
    if not image_items:
        print("No images found in provided Drive folders.")
        sys.exit(1)

    if args.sort:
        image_items.sort(key=lambda x: x[1])
    if args.shuffle:
        random.shuffle(image_items)

    if args.max_files and args.max_files > 0:
        image_items = image_items[: args.max_files]

    print(f"Found {len(image_items)} image files to process.")

    processed: Set[str] = set()
    if args.resume:
        processed = load_processed_ids(log_path)
        print(f"Resume enabled. Already processed: {len(processed)}")

    rows_to_write: List[Row] = []
    ok_count = 0
    skip_count = 0
    err_count = 0

    default_microfilm = args.microfilm_name.strip()
    if not default_microfilm and not args.microfilm_name_from_folder:
        default_microfilm = "UNKNOWN"

    # Make sure output exists early (helps if the run dies mid-way)
    ensure_excel(out_xlsx)

    for i, (file_id, rel_path) in enumerate(image_items, 1):
        if args.resume and file_id in processed:
            skip_count += 1
            continue

        try:
            local_path = cache_dir / rel_path
            if not local_path.exists():
                print(f"[{i}/{len(image_items)}] Downloading: {rel_path}")
                drive_download_file(svc, file_id, local_path)
            else:
                print(f"[{i}/{len(image_items)}] Cached: {rel_path}")

            row = process_one_image(
                local_path=local_path,
                drive_rel_path=rel_path,
                countries=countries,
                states=states,
                locations=locations,
                regions_map=regions_map,
                ioc_index=ioc_index,
                geocode_fn=geocode_fn,
                save_ocr_dir=save_ocr_dir,
                microfilm_name=default_microfilm,
                microfilm_from_folder=args.microfilm_name_from_folder,
                interactive=args.interactive,
                psm=args.psm,
                oem=args.oem,
            )
            rows_to_write.append(row)
            ok_count += 1

            append_log(log_path, {
                "file_id": file_id,
                "rel_path": rel_path,
                "status": "ok",
                "psm": args.psm,
                "oem": args.oem,
            })
            print(f"  -> OK: {local_path.name}")

            # Write in batches so work is not lost
            if len(rows_to_write) >= 25:
                append_rows_to_excel(out_xlsx, rows_to_write)
                rows_to_write = []
                print("  -> wrote batch to Excel")

        except Exception as e:
            err_count += 1
            append_log(log_path, {"file_id": file_id, "rel_path": rel_path, "status": "error", "error": str(e)})
            print(f"  -> ERROR: {rel_path} :: {e}")

    if rows_to_write:
        append_rows_to_excel(out_xlsx, rows_to_write)

    print("\nDone.")
    print(f"  OK:   {ok_count}")
    print(f"  SKIP: {skip_count}")
    print(f"  ERR:  {err_count}")
    print(f"Output: {out_xlsx}")
    print(f"Log:    {log_path}")

if __name__ == "__main__":
    main()
